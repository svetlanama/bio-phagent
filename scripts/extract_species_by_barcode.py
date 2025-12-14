#!/usr/bin/env python3
"""
Extract dominant bacterial species for specified barcodes and add to Excel metadata file.
Creates a new sheet for each barcode with all species detected.

Usage:
    python extract_species_by_barcode.py BC01
    python extract_species_by_barcode.py BC01 BC02 BC03
    python extract_species_by_barcode.py --all
"""

import argparse
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Base paths
BASE_DIR = Path(__file__).parent.parent
EXPORT_DIR = BASE_DIR / "export"
OUTPUT_DIR = BASE_DIR / "output"
BASIC_DIR = OUTPUT_DIR / "basic"
XLSX_DIR = OUTPUT_DIR / "xlsx"

# All available barcodes
ALL_BARCODES = [f"BC{i:02d}" for i in range(1, 25)]


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Extract dominant bacterial species for specified barcodes and add to Excel."
    )
    parser.add_argument(
        "barcodes",
        nargs="*",
        help="Barcodes to process (e.g., BC01 BC02 BC03)"
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Process all 24 barcodes"
    )
    return parser.parse_args()


def load_species_data():
    """Load species abundance data."""
    species_file = EXPORT_DIR / "abundances" / "table" / "abundance_species_table.csv"
    df = pd.read_csv(species_file)
    return df


def load_metadata():
    """Load barcode to farm metadata mapping."""
    mapping_file = BASIC_DIR / "barcode_farm_mapping.csv"
    df = pd.read_csv(mapping_file)
    # Create lookup dict: barcode01 -> {sample_id, farm_type, ...}
    metadata = {}
    for _, row in df.iterrows():
        metadata[row['barcode']] = {
            'sample_id': row['sample_id'],
            'farm_type': row['farm_type'],
            'farm_id': row['farm_id'],
            'sample_type': row['sample_type'],
            'oblast': row['oblast']
        }
    return metadata


def bc_to_barcode(bc):
    """Convert BC01 format to barcode01 format."""
    num = int(bc.replace("BC", ""))
    return f"barcode{num:02d}"


def barcode_to_bc(barcode):
    """Convert barcode01 format to BC01 format."""
    num = int(barcode.replace("barcode", ""))
    return f"BC{num:02d}"


def extract_species(df, barcode):
    """Extract and format species data for a given barcode."""
    barcode_col = bc_to_barcode(barcode)

    if barcode_col not in df.columns:
        print(f"  Warning: {barcode_col} not found in data")
        return None, 0

    # Filter to species with counts > 0
    data = df[df[barcode_col] > 0].copy()

    if len(data) == 0:
        print(f"  Warning: No species found for {barcode}")
        return None, 0

    # Select relevant columns
    columns_to_keep = ['species', barcode_col, 'genus', 'family', 'order', 'class', 'phylum']
    data = data[columns_to_keep].copy()

    # Rename columns
    data = data.rename(columns={
        barcode_col: 'Read Count',
        'species': 'Species',
        'genus': 'Genus',
        'family': 'Family',
        'order': 'Order',
        'class': 'Class',
        'phylum': 'Phylum'
    })

    # Calculate total reads and relative abundance
    total_reads = data['Read Count'].sum()
    data['Relative Abundance (%)'] = (data['Read Count'] / total_reads * 100).round(2)

    # Sort by read count descending
    data = data.sort_values('Read Count', ascending=False).reset_index(drop=True)

    # Add rank
    data['Rank'] = range(1, len(data) + 1)

    # Reorder columns
    final_columns = ['Rank', 'Species', 'Read Count', 'Relative Abundance (%)',
                     'Genus', 'Family', 'Order', 'Class', 'Phylum']
    data = data[final_columns]

    return data, total_reads


def add_to_excel(species_data, barcode, metadata, total_reads, wb):
    """Add species data as a new sheet to the workbook."""
    sheet_name = f"{barcode}_Species"

    # Remove existing sheet if it exists
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # Create new sheet
    ws = wb.create_sheet(sheet_name)

    # Get sample metadata
    barcode_col = bc_to_barcode(barcode)
    meta = metadata.get(barcode_col, {})
    sample_id = meta.get('sample_id', 'N/A')
    farm_type = meta.get('farm_type', 'N/A')
    farm_id = meta.get('farm_id', 'N/A')
    sample_type = meta.get('sample_type', 'N/A')
    oblast = meta.get('oblast', 'N/A')

    # Add header info
    ws['A1'] = f"{barcode} - Dominant Bacterial Species"
    ws['A2'] = f"Sample: {sample_id} | Farm: {farm_id} ({farm_type}) | Type: {sample_type} | Oblast: {oblast}"
    ws['A3'] = f"Total Reads: {total_reads:,} | Unique Species: {len(species_data)}"

    # Add empty row
    start_row = 5

    # Add column headers
    for col_idx, col_name in enumerate(species_data.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)

    # Add data rows
    for row_idx, row in enumerate(species_data.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    column_widths = {
        'A': 8,   # Rank
        'B': 35,  # Species
        'C': 12,  # Read Count
        'D': 20,  # Relative Abundance
        'E': 20,  # Genus
        'F': 25,  # Family
        'G': 25,  # Order
        'H': 25,  # Class
        'I': 20,  # Phylum
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    return sheet_name


def main():
    args = parse_args()

    # Determine which barcodes to process
    if args.all:
        barcodes = ALL_BARCODES
    elif args.barcodes:
        barcodes = [bc.upper() for bc in args.barcodes]
    else:
        print("Error: Please specify barcodes or use --all")
        print("Usage: python extract_species_by_barcode.py BC01 BC02")
        print("       python extract_species_by_barcode.py --all")
        return

    # Validate barcodes
    invalid = [bc for bc in barcodes if bc not in ALL_BARCODES]
    if invalid:
        print(f"Error: Invalid barcodes: {invalid}")
        print(f"Valid barcodes: BC01-BC24")
        return

    print(f"Extracting species data for: {', '.join(barcodes)}")
    print("=" * 50)

    # Load data
    df = load_species_data()
    metadata = load_metadata()
    print(f"Loaded {len(df)} species from abundance table")

    # Load Excel workbook
    excel_file = XLSX_DIR / "[WIP] UA_FARM_WW_CLEAN_METADATA _v1.xlsx"
    wb = load_workbook(excel_file)

    # Process each barcode
    sheets_added = []
    for barcode in barcodes:
        print(f"\nProcessing {barcode}...")

        species_data, total_reads = extract_species(df, barcode)

        if species_data is None:
            continue

        print(f"  Total reads: {total_reads:,}")
        print(f"  Unique species: {len(species_data)}")

        # Show top 5
        print(f"  Top 5 species:")
        for _, row in species_data.head(5).iterrows():
            print(f"    {row['Rank']:2}. {row['Species']}: {row['Read Count']:,} ({row['Relative Abundance (%)']}%)")

        # Add to Excel
        sheet_name = add_to_excel(species_data, barcode, metadata, total_reads, wb)
        sheets_added.append(sheet_name)

    # Save workbook
    wb.save(excel_file)

    print("\n" + "=" * 50)
    print(f"Added {len(sheets_added)} sheets to: {excel_file}")
    for sheet in sheets_added:
        print(f"  - {sheet}")
    print("Done!")


if __name__ == "__main__":
    main()
