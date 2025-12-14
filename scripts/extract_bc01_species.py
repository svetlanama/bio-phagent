#!/usr/bin/env python3
"""
Extract dominant bacterial species for BC01 and add to Excel metadata file.
Creates a new sheet with all species detected in barcode01 sample.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Base paths
BASE_DIR = Path(__file__).parent.parent
EXPORT_DIR = BASE_DIR / "export"
OUTPUT_DIR = BASE_DIR / "output"


def load_species_data():
    """Load species abundance data."""
    species_file = EXPORT_DIR / "abundances" / "table" / "abundance_species_table.csv"
    df = pd.read_csv(species_file)
    return df


def extract_bc01_species(df):
    """Extract and format BC01 species data."""
    # Filter to species with counts > 0 for barcode01
    bc01_data = df[df['barcode01'] > 0].copy()

    # Select relevant columns
    columns_to_keep = ['species', 'barcode01', 'genus', 'family', 'order', 'class', 'phylum']
    bc01_data = bc01_data[columns_to_keep].copy()

    # Rename columns
    bc01_data = bc01_data.rename(columns={
        'barcode01': 'Read Count',
        'species': 'Species',
        'genus': 'Genus',
        'family': 'Family',
        'order': 'Order',
        'class': 'Class',
        'phylum': 'Phylum'
    })

    # Calculate total reads and relative abundance
    total_reads = bc01_data['Read Count'].sum()
    bc01_data['Relative Abundance (%)'] = (bc01_data['Read Count'] / total_reads * 100).round(2)

    # Sort by read count descending
    bc01_data = bc01_data.sort_values('Read Count', ascending=False).reset_index(drop=True)

    # Add rank
    bc01_data['Rank'] = range(1, len(bc01_data) + 1)

    # Reorder columns
    final_columns = ['Rank', 'Species', 'Read Count', 'Relative Abundance (%)',
                     'Genus', 'Family', 'Order', 'Class', 'Phylum']
    bc01_data = bc01_data[final_columns]

    return bc01_data, total_reads


def add_to_excel(bc01_data, total_reads):
    """Add BC01 species data as a new sheet to the Excel file."""
    excel_file = OUTPUT_DIR / "[WIP] UA_FARM_WW_CLEAN_METADATA _v1.xlsx"

    # Load existing workbook
    wb = load_workbook(excel_file)

    # Remove existing BC01_Species sheet if it exists
    if "BC01_Species" in wb.sheetnames:
        del wb["BC01_Species"]

    # Create new sheet
    ws = wb.create_sheet("BC01_Species")

    # Add header info
    ws['A1'] = "BC01 - Dominant Bacterial Species"
    ws['A2'] = f"Sample: 1.1. | Farm: Poultry #1 | Type: Native | Oblast: Zaporizhzhia"
    ws['A3'] = f"Total Reads: {total_reads:,} | Unique Species: {len(bc01_data)}"

    # Add empty row
    start_row = 5

    # Add column headers
    for col_idx, col_name in enumerate(bc01_data.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)

    # Add data rows
    for row_idx, row in enumerate(bc01_data.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    column_widths = {
        'A': 8,   # Rank
        'B': 35,  # Species
        'C': 12,  # Read Count
        'D': 20,  # Relative Abundance
        'E': 20,  # Genus
        'F': 25,  # Family
        'G': 25,  # Order
        'H': 25,  # Class
        'I': 20,  # Phylum
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save workbook
    wb.save(excel_file)
    print(f"Added BC01_Species sheet to: {excel_file}")


def main():
    print("Extracting BC01 species data...")
    print("=" * 50)

    # Load data
    df = load_species_data()
    print(f"Loaded {len(df)} species from abundance table")

    # Extract BC01 species
    bc01_data, total_reads = extract_bc01_species(df)
    print(f"\nBC01 Summary:")
    print(f"  Total reads: {total_reads:,}")
    print(f"  Unique species: {len(bc01_data)}")

    # Show top 10
    print(f"\nTop 10 Species:")
    for _, row in bc01_data.head(10).iterrows():
        print(f"  {row['Rank']:2}. {row['Species']}: {row['Read Count']:,} ({row['Relative Abundance (%)']}%)")

    # Add to Excel
    add_to_excel(bc01_data, total_reads)

    print("\n" + "=" * 50)
    print("Done!")


if __name__ == "__main__":
    main()
